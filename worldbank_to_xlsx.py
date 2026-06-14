from __future__ import annotations

import asyncio
import base64
import csv
import hashlib
import json
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode

import aiohttp
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm

WDS_URL = "https://search.worldbank.org/api/v3/wds"
PROJECTS_URL = "https://search.worldbank.org/api/v3/projects"
WDS_PARAMS = {
	"fl": "projectid",
	"rows": "1000",
	"docty_exact": "Procurement Plan",
	"count_exact": "",
	"sectr_exact": "",
}
FIELDNAME_WIDTHS = {
	"Country": 30,
	"Project ID": 10,
	"Project Title": 40,
	"Team Leader": 40,
	"Effective Date": 20,
	"Closing Date": 20,
	"Total Project Cost": 20,
	"Implementing Agency": 40,
	"Procurement Plan Link": 40,
	"Activity Reference No. / Description": 40,
	"Method": 30,
	"Market Approach": 30,
	"Estimated Amount (US$)": 30,
	"Process Status": 30,
	"Expression of Interest Notice": 30,
}
FIELDNAMES = list(FIELDNAME_WIDTHS.keys())
DATE_FIELDNAMES = {"Effective Date", "Closing Date", "Expression of Interest Notice"}
NA_DATE_FIELDNAMES = {"Effective Date", "Closing Date"}
CLOSING_DATE_FIELDNAME = "Closing Date"
CLOSING_DATE_CUTOFF = datetime(2026, 7, 1)
BACKGROUND_COLOR = "F1A983"
HYPERLINK_COLOR = "0070C0"
DATE_FORMAT = "DD/MM/YYYY"
WIDTH_OFFSET = 0.71
BASE_DIR = Path.home() / ".cache/scripts"
CACHE_DIR = BASE_DIR / "cache"
OUTPUT_CSV = "projects.csv"
OUTPUT_XLSX = "projects.xlsx"


def generate_cache_id() -> str:
	seed = f"{WDS_PARAMS['count_exact']}|{WDS_PARAMS['sectr_exact']}"
	hashed = hashlib.sha256(seed.encode("utf-8")).digest()
	encoded = base64.urlsafe_b64encode(hashed).decode("ascii")
	return encoded[:11]


CACHE_ID = generate_cache_id()
CACHE_ID_DIR = CACHE_DIR / CACHE_ID
CSV_CACHE_DIR = CACHE_ID_DIR / "csv"
JSON_CACHE_DIR = CACHE_ID_DIR / "json"
PDF_CACHE_DIR = CACHE_ID_DIR / "pdf"
XLSX_CACHE_DIR = CACHE_ID_DIR / "xlsx"
CACHE_DIRS = [
	CACHE_ID_DIR,
	CSV_CACHE_DIR,
	JSON_CACHE_DIR,
	PDF_CACHE_DIR,
	XLSX_CACHE_DIR,
]


def cache_init() -> None:
	for cache_dir in CACHE_DIRS:
		cache_dir.mkdir(parents=True, exist_ok=True)


async def fetch(session: aiohttp.ClientSession, url: str) -> bytes | None:
	async with session.get(url) as resp:
		if resp.status == 200:
			return await resp.read()
	return None


async def fetch_and_cache_json(
	session: aiohttp.ClientSession, url: str, cache_path: Path
) -> dict:
	if cache_path.exists():
		with open(cache_path, "r", encoding="utf-8") as file:
			return json.load(file)
	data = await fetch(session, url)
	if data is None:
		return {}
	obj = json.loads(data.decode("utf-8"))
	with open(cache_path, "w", encoding="utf-8") as file:
		json.dump(obj, file, ensure_ascii=False, indent="\t")
	return obj


async def extract_project_ids(
	session: aiohttp.ClientSession, offset: int
) -> tuple[set[str], int]:
	cache_path = JSON_CACHE_DIR / f"wds_{offset}.json"
	params = WDS_PARAMS.copy()
	params["os"] = str(offset)
	url = f"{WDS_URL}?{urlencode(params)}"
	obj = await fetch_and_cache_json(session, url, cache_path)
	project_ids = set()
	total = obj.get("total", 0)
	documents = obj.get("documents", {})
	for key, value in documents.items():
		if key == "facets":
			continue
		project_id = value.get("projectid")
		project_ids.add(project_id)
	return project_ids, total


async def extract_all_project_ids(session: aiohttp.ClientSession) -> list[str]:
	project_ids, total = await extract_project_ids(session, 0)
	all_project_ids = set(project_ids)
	offsets = range(1000, total, 1000)
	tasks = [extract_project_ids(session, offset) for offset in offsets]
	pbar = tqdm(total=len(tasks), desc="Pages")
	for task in asyncio.as_completed(tasks):
		project_ids, _ = await task
		all_project_ids.update(project_ids)
		pbar.update()
	pbar.close()
	return sorted(all_project_ids)


def format_cost(us_cost: str) -> str:
	if not us_cost:
		return ""
	cost = float(us_cost)
	eu_cost = f"{cost:,.2f}".replace(",", " ").replace(".", ",")
	return f"$ {eu_cost}"


def extract_team_leader_names(project: dict) -> str:
	teamleadnames = project.get("teamleadname", [])
	return ", ".join(name.strip() for name in teamleadnames if name.strip())


def parse_project(project_id: str, obj: dict) -> dict | None:
	project = obj.get("projects", {}).get(project_id)
	if not project:
		return None
	return {
		"Country": project.get("countryshortname", ""),
		"Project ID": project_id,
		"Project Title": project.get("project_name", ""),
		"Team Leader": extract_team_leader_names(project),
		"Effective Date": project.get("loan_effective_date", ""),
		"Closing Date": project.get("closingdate", ""),
		"Total Project Cost": format_cost(project.get("lendprojectcost", "")),
		"Implementing Agency": project.get("impagency", ""),
	}


async def extract_project(
	session: aiohttp.ClientSession, project_id: str
) -> tuple[str, dict | None]:
	cache_path = JSON_CACHE_DIR / f"projects_{project_id}.json"
	params = {"fl": "*", "id": project_id}
	url = f"{PROJECTS_URL}?{urlencode(params)}"
	obj = await fetch_and_cache_json(session, url, cache_path)
	return project_id, parse_project(project_id, obj)


async def extract_projects(
	session: aiohttp.ClientSession, project_ids: list[str]
) -> dict[str, dict]:
	projects = {}
	tasks = [extract_project(session, project_id) for project_id in project_ids]
	pbar = tqdm(total=len(tasks), desc="Projects")
	for task in asyncio.as_completed(tasks):
		project_id, project = await task
		if project:
			projects[project_id] = project
		pbar.update()
	pbar.close()
	return projects


async def extract_procurement_plan_link(
	session: aiohttp.ClientSession, project_id: str
) -> tuple[str, str]:
	cache_path = JSON_CACHE_DIR / f"wds_{project_id}.json"
	params = {
		"fl": "pdfurl",
		"rows": "1",
		"docty_exact": "Procurement Plan",
		"proid": project_id,
	}
	url = f"{WDS_URL}?{urlencode(params)}"
	obj = await fetch_and_cache_json(session, url, cache_path)
	pdf_url = ""
	documents = obj.get("documents", {})
	for key, value in documents.items():
		if key == "facets":
			continue
		pdf_url = value.get("pdfurl", "")
	return project_id, pdf_url


async def extract_procurement_plan_links(
	session: aiohttp.ClientSession, project_ids: list[str]
) -> dict[str, str]:
	pdf_urls = {}
	tasks = [
		extract_procurement_plan_link(session, project_id) for project_id in project_ids
	]
	pbar = tqdm(total=len(tasks), desc="Plans")
	for task in asyncio.as_completed(tasks):
		project_id, pdf_url = await task
		if pdf_url:
			pdf_urls[project_id] = pdf_url
		pbar.update()
	pbar.close()
	return pdf_urls


def is_after_cutoff(iso_closing_date: str) -> bool:
	if not iso_closing_date:
		return True
	year, month, day = iso_closing_date.split("-")
	closing_date = datetime(int(year), int(month), int(day))
	return closing_date >= CLOSING_DATE_CUTOFF


def build_project_rows(
	projects: dict[str, dict], pdf_urls: dict[str, str]
) -> list[dict]:
	rows = []
	for project_id in sorted(projects.keys()):
		project = projects[project_id]
		pdf_url = ""
		if is_after_cutoff(project.get("Closing Date", "")):
			pdf_url = pdf_urls.get(project_id, "")
		project["Procurement Plan Link"] = pdf_url
		rows.append(project)
	rows.sort(key=lambda row: (row["Country"], row["Project ID"]))
	return rows


def write_csv(rows: list[dict]) -> None:
	with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as file:
		writer = csv.DictWriter(file, fieldnames=FIELDNAMES)
		writer.writeheader()
		writer.writerows(rows)


def parse_date(iso_date: str, allow_na: bool) -> datetime | str:
	if not iso_date:
		return "N/A" if allow_na else ""
	year, month, day = iso_date.split("-")
	closing_date = datetime(int(year), int(month), int(day))
	return closing_date


def format_row_values(row: dict, headers: list[str], is_project: bool) -> list:
	values = []
	for header in headers:
		value = row.get(header, "")
		if header in DATE_FIELDNAMES:
			allow_na = header in NA_DATE_FIELDNAMES and is_project
			value = parse_date(value, allow_na)
		values.append(value)
	return values


def apply_width(worksheet: Worksheet, headers: list[str]) -> None:
	for col_idx, header in enumerate(headers, start=1):
		width = FIELDNAME_WIDTHS.get(header) + WIDTH_OFFSET
		worksheet.column_dimensions[get_column_letter(col_idx)].width = width


def apply_alignment(worksheet: Worksheet) -> None:
	alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
	for row in worksheet.iter_rows():
		for cell in row:
			cell.alignment = alignment


def apply_bold_header(worksheet: Worksheet) -> None:
	bold_font = Font(bold=True)
	for cell in worksheet[1]:
		cell.font = bold_font


def apply_date_format(worksheet: Worksheet, headers: list[str]) -> None:
	for col_idx, header in enumerate(headers, start=1):
		if header not in DATE_FIELDNAMES:
			continue
		for row_idx in range(2, worksheet.max_row + 1):
			cell = worksheet.cell(row=row_idx, column=col_idx)
			if isinstance(cell.value, datetime):
				cell.number_format = DATE_FORMAT


def apply_hyperlink(worksheet: Worksheet) -> None:
	hyperlink_font = Font(color=HYPERLINK_COLOR, underline="single")
	for row in worksheet.iter_rows(min_row=2):
		for cell in row:
			if isinstance(cell.value, str) and cell.value.startswith(
				("http://", "https://")
			):
				cell.hyperlink = cell.value
				cell.font = hyperlink_font


def apply_conditional_formatting(worksheet: Worksheet, headers: list[str]) -> None:
	col_idx = headers.index(CLOSING_DATE_FIELDNAME) + 1
	closing_date_column_letter = get_column_letter(col_idx)
	last_column_letter = get_column_letter(len(headers))
	formula = f"AND(NOT(ISBLANK(${closing_date_column_letter}2)), ${closing_date_column_letter}2<DATE({CLOSING_DATE_CUTOFF.strftime("%Y,%m,%d")}))"
	fill = PatternFill(end_color=BACKGROUND_COLOR)
	worksheet.conditional_formatting.add(
		f"A2:{last_column_letter}{worksheet.max_row}",
		FormulaRule(formula=[formula], fill=fill),
	)


def write_excel(rows: list[dict], headers: list[str]) -> None:
	workbook = Workbook()
	worksheet = workbook.active
	worksheet.title = "Projects"
	worksheet.append(headers)
	for row in rows:
		is_project_row = bool(row.get("Project ID", "").strip())
		values = format_row_values(row, headers, is_project_row)
		worksheet.append(values)
	apply_width(worksheet, headers)
	apply_alignment(worksheet)
	apply_bold_header(worksheet)
	apply_date_format(worksheet, headers)
	apply_hyperlink(worksheet)
	apply_conditional_formatting(worksheet, headers)
	workbook.save(OUTPUT_XLSX)


async def download_pdf(session: aiohttp.ClientSession, pdf_url: str) -> str | None:
	filename = pdf_url.split("/")[-1]
	cache_path = PDF_CACHE_DIR / filename
	if cache_path.exists():
		return cache_path.name
	content = await fetch(session, pdf_url)
	with open(cache_path, "wb") as file:
		file.write(content)
	return cache_path.name


async def download_pdfs(session: aiohttp.ClientSession, rows: list[dict]) -> None:
	tasks = []
	for row in rows:
		pdf_url = row.get("Procurement Plan Link", "")
		if pdf_url:
			tasks.append(download_pdf(session, pdf_url))
	if not tasks:
		print("No PDFs to download")
		return
	pbar = tqdm(total=len(tasks), desc="PDFs")
	for task in asyncio.as_completed(tasks):
		await task
		pbar.update()
	pbar.close()


async def main() -> None:
	cache_init()
	async with aiohttp.ClientSession() as session:
		project_ids = await extract_all_project_ids(session)
		projects = await extract_projects(session, project_ids)
		pdf_urls = await extract_procurement_plan_links(session, project_ids)
		rows = build_project_rows(projects, pdf_urls)
		write_csv(rows)
		write_excel(rows, FIELDNAMES)
		await download_pdfs(session, rows)


if __name__ == "__main__":
	asyncio.run(main())
